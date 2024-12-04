import streamlit as st
import openpyxl as op

# Define the scaler function here or import it from your existing script.
import openpyxl as op

def scaler(workbook1):
    from datetime import datetime

    def col_letter(n):
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
    
        return result

    sheet = workbook1.active

    total_rows = sheet.max_row
    total_columns = sheet.max_column

    total_students=total_rows-1
    sorted_data=[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        temp=[]
        for i in range(len(row)):
            temp.append(row[i])
        sorted_data.append(temp)

    sorted_data = sorted(sorted_data, key=lambda x: x[2],reverse=True)

    
    workbook2 = op.Workbook()
    sheet2=workbook2.active
    datasheet={
        "AA":{
            "min":"",
            "max":"",
        },
        "AB":{
            "min":"",
            "max":"",
        },
        "BB":{
            "min":"",
            "max":"",
        },
        "BC":{
            "min":"",
            "max":"",
        },
        "CC":{
            "min":"",
            "max":"",
        },
        "CD":{
            "min":"",
            "max":"",
        },
        "DD":{
            "min":"",
            "max":"",
        },
        "F":{
            "min":"",
            "max":"",
        },
    }
    sheet2.cell(row=17, column= 1, value="Roll")
    sheet2.cell(row=17, column= 2, value="Name")
    sheet2.cell(row=17, column= 3, value="Total")
    sheet2.cell(row=17, column= 4, value="Grade")
    sheet2.cell(row=17, column= 5, value="Scaled")
    for row in range(18,total_rows+17):
        for col in range(0,total_columns):
            val=sorted_data[row-18][col]
            # print(val)
            sheet2.cell(row=row, column=col+1, value=val)
            if(col==3):
                datasheet[val]['min']=f"=C{row}"
                if(datasheet[val]['max']==""):
                    datasheet[val]['max']=f"=C{row}"

    today = datetime.now()
    formatted_date = today.strftime("%b-%d")
    
    sheet2['A1']="Subject Code"
    sheet2['A2']="Grade"
    sheet2['A3']="AA"
    sheet2['A4']="AB"
    sheet2['A5']="BB"
    sheet2['A6']="BC"
    sheet2['A7']="CC"
    sheet2['A8']="CD"
    sheet2['A9']="DD"
    sheet2['A10']="F"
    sheet2['B2']="a"
    sheet2['B3']="91"
    sheet2['B4']="81"
    sheet2['B5']="71"
    sheet2['B6']="61"
    sheet2['B7']="51"
    sheet2['B8']="41"
    sheet2['B9']="31"
    sheet2['B10']="0"
    sheet2['C2']="b"
    sheet2['C3']="100"
    sheet2['C4']="90"
    sheet2['C5']="80"
    sheet2['C6']="70"
    sheet2['C7']="60"
    sheet2['C8']="50"
    sheet2['C9']="40"
    sheet2['C10']="30"
    sheet2['D2']="min(x)"
    sheet2['D3']=datasheet['AA']['min']
    sheet2['D4']=datasheet['AB']['min']
    sheet2['D5']=datasheet['BB']['min']
    sheet2['D6']=datasheet['BC']['min']
    sheet2['D7']=datasheet['CC']['min']
    sheet2['D8']=datasheet['CD']['min']
    sheet2['D9']=datasheet['CC']['min']
    sheet2['D10']=datasheet['F']['min']
    sheet2['E1']=formatted_date
    sheet2['E2']="max(x)"
    sheet2['E3']=datasheet['AA']['max']
    sheet2['E4']=datasheet['AB']['max']
    sheet2['E5']=datasheet['BB']['max']
    sheet2['E6']=datasheet['BC']['max']
    sheet2['E7']=datasheet['CC']['max']
    sheet2['E8']=datasheet['CD']['max']
    sheet2['E9']=datasheet['CC']['max']
    sheet2['E10']=datasheet['F']['max']
    sheet2['F2']="Grade"
    sheet2['F3']="AA"
    sheet2['F4']="AB"
    sheet2['F5']="BB"
    sheet2['F6']="BC"
    sheet2['F7']="CC"
    sheet2['F8']="CD"
    sheet2['F9']="DD"
    sheet2['F10']="F"
    sheet2['F11']="TOTAL "
    sheet2['G2']="Count"
    sheet2['G3']=f'=COUNTIF(D18:D{17+total_students},"AA")'
    sheet2['G4']=f'=COUNTIF(D18:D{17+total_students},"AB")'
    sheet2['G5']=f'=COUNTIF(D18:D{17+total_students},"BB")'
    sheet2['G6']=f'=COUNTIF(D18:D{17+total_students},"BC")'
    sheet2['G7']=f'=COUNTIF(D18:D{17+total_students},"CC")'
    sheet2['G8']=f'=COUNTIF(D18:D{17+total_students},"CD")'
    sheet2['G9']=f'=COUNTIF(D18:D{17+total_students},"DD")'
    sheet2['G10']=f'=COUNTIF(D18:D{17+total_students},"F")'
    sheet2['G11']=f'=sum(G3:G10)'
    sheet2['I2']="Grade"
    sheet2['I3']="AA"
    sheet2['I4']="AB"
    sheet2['I5']="BB"
    sheet2['I6']="BC"
    sheet2['I7']="CC"
    sheet2['I8']="CD"
    sheet2['I9']="DD"
    sheet2['I10']="F"
    sheet2['J1']="Stud Count"
    sheet2['J2']="IAPC"
    sheet2['J3']="5"
    sheet2['J4']="15"
    sheet2['J5']="25"
    sheet2['J6']="30"
    sheet2['J7']="15"
    sheet2['J8']="5"
    sheet2['J9']="5"
    sheet2['J10']="0"
    sheet2['K1']=f"=COUNTA(A18:A{17+total_students})"
    sheet2['K2']="IAPC Count"
    sheet2['K3'].value='='+f"CEILING(J3/100*$K$1,1)"
    sheet2['K4']="=CEILING(J4/100*$K$1,1)"
    sheet2['K5']="=CEILING(J5/100*$K$1,1)"
    sheet2['K6']="=CEILING(J6/100*$K$1,1)"
    sheet2['K7']="=CEILING(J7/100*$K$1,1)"
    sheet2['K8']="=CEILING(J8/100*$K$1,1)"
    sheet2['K9']="=CEILING(J9/100*$K$1,1)"
    sheet2['K10']="=CEILING(J10/100*$K$1,1)"
    sheet2['L2']="Diff"
    sheet2['L3']="=(G3-K3)"
    sheet2['L4']=f"=(G4-K4)"
    sheet2['L5']="=G5-K5"
    sheet2['L6']="=G6-K6"
    sheet2['L7']="=G7-K7"
    sheet2['L8']="=G8-K8"
    sheet2['L9']="=G9-K9"
    sheet2['L10']="=G10-K10"
    from openpyxl.drawing.image import Image
    img=Image("image.jpg")
    sheet2.add_image(img,"B12")
    # worksheet2.formula_attributes[cell.coordinate] = {'t': 'array', 'ref': f"{cell.coordinate}:{cell.coordinate}"}
    formula={
        "AA":"=9*((CX-$D$3)/($E$3-$D$3))+$B$3",
        "AB":"=9*((CX-$D$4)/($E$4-$D$4))+$B$4",
        "BB":"=9*((CX-$D$5)/($E$5-$D$5))+$B$5",
        "BC":"=9*((CX-$D$6)/($E$6-$D$6))+$B$6",
        "CC":"=9*((CX-$D$7)/($E$7-$D$7))+$B$7",
        "CD":"=9*((CX-$D$8)/($E$8-$D$8))+$B$8",
        "DD":"=9*((CX-$D$9)/($E$9-$D$9))+$B$9",
        "F":"=9*((CX-$D$10)/($E$10-$D$10))+$B$10"
    }
    for i in range(18,18+total_students):
        sheet2["E"+str(i)]=formula[sheet2['D'+str(i)].value].replace('X',str(i))
    from openpyxl.styles import Font,Border, Side
    new_font = Font(size=15)  # Change the size as needed

# Apply the font size to all cells in the workbook
    for sheet in workbook2.worksheets:  # Iterate through all sheets
        for row in sheet.iter_rows():  # Iterate through all rows
            for cell in row:  # Iterate through all cells in a row
                if cell.value is not None:  # Apply only to non-empty cells
                    cell.font = new_font
    sheet2.column_dimensions['A'].width = 15
    sheet2.column_dimensions['B'].width = 37
    sheet2.column_dimensions['C'].width = 17
    sheet2.column_dimensions['D'].width = 10
    sheet2.column_dimensions['E'].width = 17
    sheet2.column_dimensions['F'].width = 12
    sheet2.column_dimensions['G'].width = 8
    sheet2.column_dimensions['H'].width = 4
    sheet2.column_dimensions['J'].width = 15
    sheet2.column_dimensions['K'].width = 12

    thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
    )

    for row in sheet2.iter_rows(min_row=1, max_row=11, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
    for row in sheet2.iter_rows(min_row=1, max_row=10, min_col=9, max_col=12):
        for cell in row:
            cell.border = thin_border
    for row in sheet2.iter_rows(min_row=17, max_row=17+total_students, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

   
    return workbook2

st.title("Excel Scaler Application")
st.write("Upload an Excel file to process and download the output.")

# File uploader
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

if uploaded_file is not None:
    # Load the uploaded workbook
    workbook = op.load_workbook(uploaded_file)

    # Process the workbook using scaler function
    st.write("Processing the file...")
    output_workbook = scaler(workbook)

    # Save the output to a BytesIO buffer
    from io import BytesIO
    output_buffer = BytesIO()
    output_workbook.save(output_buffer)
    output_buffer.seek(0)

    # Provide a download button for the user
    st.download_button(
        label="Download Processed Excel File",
        data=output_buffer,
        file_name="Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.success("File processed successfully!")