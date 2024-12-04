import os
from flask import Flask, render_template, request, send_file, redirect, url_for
import openpyxl

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed'

# Ensure upload and processed directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)
def add_summary_table(ws, num_students, old_iapc_reco_percentages, grade_counts, grade_round_counts, start_col):
    """Writes the summary table to a given worksheet at the specified starting column."""
    summary_data = [
        ["Total Students", num_students],
        ["grade", "old iapc reco", "Counts", "Round", "Count verified"]
    ]

    # Adding each grade's row with old iapc reco calculations
    for grade in old_iapc_reco_percentages.keys():
        row_data = [
            grade,
            old_iapc_reco_percentages[grade],  # "old iapc reco" percentage
            grade_counts[grade],               # Calculated count
            grade_round_counts[grade],         # Rounded count
            grade_round_counts[grade]          # For "Count verified" column
        ]
        summary_data.append(row_data)

    # Append additional rows for unused grades
    extra_grades = ['F', 'I', 'PP', 'NP']
    for extra_grade in extra_grades:
        summary_data.append([extra_grade, 0, 0, 0, 0])

    # Write the summary data to the worksheet starting from the specified column
    for i, row_data in enumerate(summary_data):
        for j, value in enumerate(row_data):
            ws.cell(row=i+1, column=start_col + j, value=value)

def process_excel_file(input_path, output_path):
    # Open the workbook and select the active worksheet
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Detect assessment columns by reading the header row
    headers = [ws.cell(row=1, column=i).value for i in range(3, ws.max_column + 1)]
    max_marks = [ws.cell(row=2, column=i).value for i in range(3, ws.max_column + 1)]
    weightage = [ws.cell(row=3, column=i).value for i in range(3, ws.max_column + 1)]

    # Calculate the number of students
    num_students = ws.max_row - 3
    totals = []

    # Calculate grand total for each student based on variable columns
    for row in range(4, ws.max_row + 1):
        total_score = 0
        for i, header in enumerate(headers):
            score = ws.cell(row=row, column=i + 3).value
            max_score = max_marks[i]
            weight = weightage[i]
            if score is not None and max_score is not None and weight is not None:
                total_score += (score / max_score * weight)
        
        roll_no = ws.cell(row=row, column=1).value
        totals.append((row, roll_no, total_score))

    # Sort students by grand total in descending order
    totals.sort(key=lambda x: x[2], reverse=True)

    # Grade allocation as per predefined percentages
    old_iapc_reco_percentages = {'AA': 5, 'AB': 15, 'BB': 25, 'BC': 30, 'CC': 15, 'CD': 5, 'DD': 5}
    grade_counts = {grade: num_students * (percent / 100) for grade, percent in old_iapc_reco_percentages.items()}
    grade_round_counts = {grade: round(count) for grade, count in grade_counts.items()}
    
    # Assign grades
    grades = []
    current_index = 0
    for grade, count in grade_round_counts.items():
        for _ in range(count):
            if current_index < len(totals):
                row, roll_no, total = totals[current_index]
                grades.append((row, grade))
                current_index += 1

    mxCol = ws.max_column
    # Write grand total and grade back to the worksheet
    ws.cell(row=1, column=mxCol + 1, value="Grand Total/100")
    ws.cell(row=1, column=mxCol + 2, value="Grade")
    # print(totals)
    for row, total in [(t[0], t[2]) for t in totals]:
        ws.cell(row=row, column=mxCol + 1, value=total)
    for row, grade in grades:
        ws.cell(row=row, column=mxCol + 2, value=grade)

    marks_sorted_data = [
        (ws.cell(row=row, column=1).value,  # Roll number
         ws.cell(row=row, column=2).value,  # Name
         *[ws.cell(row=row, column=i).value for i in range(3, ws.max_column + 1)])
        for row in range(4, ws.max_row + 1)
    ]
    marks_sorted_data.sort(key=lambda x: x[-2], reverse = True)  # Sort by roll number

    # Clear existing data and populate sorted data in the duplicate sheet
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col, value=None)
    for i, row_data in enumerate(marks_sorted_data, start=4):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)
    
    ws.title = "Sheet1_Mark_Sorted"

    # Create a copy of the worksheet for roll number sorting
    ws_roll_sorted = wb.copy_worksheet(ws)
    ws_roll_sorted.title = "Sheet2_Roll_Sorted"

    # Sort ws_roll_sorted based on roll number (column 1)
    roll_sorted_data = [
        (ws_roll_sorted.cell(row=row, column=1).value,  # Roll number
         ws_roll_sorted.cell(row=row, column=2).value,  # Name
         *[ws_roll_sorted.cell(row=row, column=i).value for i in range(3, ws_roll_sorted.max_column + 1)])
        for row in range(4, ws_roll_sorted.max_row + 1)
    ]
    roll_sorted_data.sort(key=lambda x: x[0])  # Sort by roll number

    # Clear existing data and populate sorted data in the duplicate sheet
    for row in range(4, ws_roll_sorted.max_row + 1):
        for col in range(1, ws_roll_sorted.max_column + 1):
            ws_roll_sorted.cell(row=row, column=col, value=None)
    for i, row_data in enumerate(roll_sorted_data, start=4):
        for j, value in enumerate(row_data, start=1):
            ws_roll_sorted.cell(row=i, column=j, value=value)

    # Add summary table to "Sorted by Roll No" sheet
    add_summary_table(ws_roll_sorted, num_students, old_iapc_reco_percentages, grade_counts, grade_round_counts, mxCol + 4)
    # Add summary table to "Sorted by Marks" sheet
    add_summary_table(ws, num_students, old_iapc_reco_percentages, grade_counts, grade_round_counts, mxCol + 4)

    # Save the workbook with both sheets
    wb.save(output_path)
    print("Output Excel file with sorted sheets generated:", output_path)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Check if the POST request has the file part
        if 'file' not in request.files:
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        
        # Save the uploaded file
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(input_path)
        
        # Define output file path
        output_pathname = "roshan" + file.filename
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_pathname)
        
        # Process the Excel file
        process_excel_file(input_path, output_path)
        
        # Provide the processed file for download
        return redirect(url_for("download_file", filename=output_pathname))
    
    return render_template("index.html")

@app.route("/download/<filename>")
def download_file(filename):
    return send_file(os.path.join(app.config['PROCESSED_FOLDER'], filename), as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
