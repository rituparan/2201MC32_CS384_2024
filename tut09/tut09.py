import pandas as pd
import ast
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Function to read the dates from the file
def read_dates(file_path):
    classes_taken_dates, classes_missed_dates, exams_dates = [], [], []

    # Open the file and read the dates
    with open(file_path, 'r') as f:
        for line in f:
            if "classes_taken_dates" in line:
                classes_taken_dates = ast.literal_eval(line.split('=')[1].strip())
            elif "classes_missed_dates" in line:
                classes_missed_dates = ast.literal_eval(line.split('=')[1].strip())
            elif "exams_dates" in line:
                exams_dates = ast.literal_eval(line.split('=')[1].strip())

    return classes_taken_dates, classes_missed_dates, exams_dates

# Helper function to extract the date from timestamp in DD/MM/YYYY HH:MM:SS format
def extract_date(timestamp):
    # Assuming the timestamp is in the format "DD/MM/YYYY HH:MM:SS"
    return datetime.strptime(timestamp, "%d/%m/%Y %H:%M:%S").strftime("%d/%m/%Y")

# Function to process the attendance and generate the Excel file
def process_attendance(attendance_file, stud_list_file, dates_file):
    # Load attendance data and student list
    df_attendance = pd.read_csv(attendance_file)
    with open(stud_list_file, 'r') as f:
        students = {line.split()[0]: ' '.join(line.split()[1:]) for line in f}

    # Read dates from the 'python_dates.txt' file
    classes_taken_dates, classes_missed_dates, exams_dates = read_dates(dates_file)

    # Initialize the Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Record"

    # Set headers
    headers = ["Roll No & Name"] + classes_taken_dates + ["Total Dates", "Total Attendance Marked", "Total Attendance Allowed", "Proxy Given"]
    ws.append(headers)

    # Create fills for attendance status
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red if > 2
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for 1
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for 2
    no_fill = PatternFill()  # No fill for 0 (Absent)

    # Get the total number of classes conducted
    total_classes_conducted = len(classes_taken_dates)

    # Populate the worksheet with student data
    for roll, name in students.items():
        row = [f"{roll} {name}"]
        total_attendance_marked = 0
        proxy_given = 0  # Initialize proxy counter
        days_attended = 0  # Track number of days attended

        for date in classes_taken_dates:
            # Extract date from attendance timestamp and compare with the class date
            attendance_records = df_attendance[(df_attendance['Roll'].str.contains(roll)) & 
                                               (df_attendance['Timestamp'].apply(extract_date) == date)]
            if len(attendance_records) == 0:
                row.append(0)  # Absent
            elif len(attendance_records) == 1:
                row.append(1)  # Partial
                total_attendance_marked += 1
                proxy_given += 1  # Count this as proxy
                days_attended += 1  # Count this day as attended
            elif len(attendance_records) == 2:
                row.append(2)  # Full attendance, no proxy
                total_attendance_marked += 2
                days_attended += 1  # Count this day as attended
            elif len(attendance_records) > 2:
                row.append(3)  # Proxy if attendance > 2 (unexpected case)
                total_attendance_marked += 3
                proxy_given += 1  # Count this as proxy
                days_attended += 1  # Count this day as attended

        # actual days attended (days_attended * 2)
        total_attendance_allowed = days_attended * 2

        # Append calculated fields to the row
        row.append(total_classes_conducted)  # Total number of classes 
        row.append(total_attendance_marked)
        row.append(total_attendance_allowed)
        row.append(proxy_given)  # Proxy count is added 

        ws.append(row)

    # Applying colors 
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=len(classes_taken_dates) + 1):
        for cell in row:
            if cell.value == 0:
                cell.fill = no_fill  # No attendance
            elif cell.value == 1:
                cell.fill = yellow_fill  # Partial attendance
            elif cell.value == 2:
                cell.fill = green_fill  # Full attendance
            elif cell.value > 2:
                cell.fill = red_fill  

    # Save the workbook to an Excel file
    wb.save("output_excel_updated.xlsx")

# Main function
def main():
    attendance_file = 'input_attendance.csv'
    stud_list_file = 'stud_list.txt'
    dates_file = 'python_dates.txt'  
    process_attendance(attendance_file, stud_list_file, dates_file)

# main function
if __name__ == "__main__":
    main()
